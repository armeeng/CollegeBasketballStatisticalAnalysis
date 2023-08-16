import openpyxl
import pandas as pd
import datetime

# Get the current time
now = datetime.datetime.now()

pd.set_option('display.max_rows', None) # So you can view the whole table
pd.set_option('display.max_columns', None) # So you can view the whole table

workbook = openpyxl.load_workbook('outputexcel.xlsx')
sheet = workbook.worksheets[0]
sheet4 = workbook.worksheets[4]

Record = pd.DataFrame({"Win Reg": 0, "Loss Reg": 0, "Reg %": 0, "Win Mult": 0, "Loss Mult": 0, "Mult %": 0, "Me Win": 0, "Me Loss": 0, "Me %": 0, "Odds Win": 0, "Odds Loss": 0, "Odds %": 0}, index=[0])
#Record = pd.read_excel('outputexcel.xlsx', sheet_name='Record')

for row in sheet:
    values = [cell.value for cell in row]
    if values[4] == None:
        break
    if values[4] == 1 and values[2] > 0:
        Record['Win Reg'] += 1
    if values[4] == 1 and values[2] < 0:
        Record["Loss Reg"] += 1
    if values[4] == 2 and values[2] > 0:
        Record["Loss Reg"] += 1
    if values[4] == 2 and values[2] < 0:
        Record['Win Reg'] += 1
    if values[4] == 1 and values[3] > 0:
        Record['Win Mult'] += 1
    if values[4] == 1 and values[3] < 0:
        Record["Loss Mult"] += 1
    if values[4] == 2 and values[3] > 0:
        Record["Loss Mult"] += 1
    if values[4] == 2 and values[3] < 0:
        Record['Win Mult'] += 1
    if values[4] == values[12]:
        Record['Me Win'] += 1
    if values[4] != values[12]:
        Record['Me Loss'] += 1
    if values[4] == values[13]:
        Record['Odds Win'] += 1
    if values[4] != values[13]:
        Record['Odds Loss'] += 1


Record['Reg %'] = (Record['Win Reg']) / (Record['Win Reg'] + Record['Loss Reg'])
Record['Mult %'] = (Record['Win Mult']) / (Record['Win Mult'] + Record['Loss Mult'])
Record['Me %'] = (Record['Me Win']) / (Record['Me Win'] + Record['Me Loss'])
Record['Odds %'] = (Record['Odds Win']) / (Record['Odds Win'] + Record['Odds Loss'])

# Find the next empty row in the sheet
next_row = sheet4.max_row + 1

# Delete the empty rows in the sheet
for row in range(4, next_row):
    if sheet4.cell(row=row, column=1).value is None:
        sheet4.delete_rows(row)

# Find the next empty row in the sheet
next_row = sheet4.max_row + 1

# Write the headers to the first row
sheet4.cell(row=3, column=1).value = "Win Reg"
sheet4.cell(row=3, column=2).value = "Loss Reg"
sheet4.cell(row=3, column=3).value = "Reg %"
sheet4.cell(row=3, column=4).value = "Win Mult"
sheet4.cell(row=3, column=5).value = "Loss Mult"
sheet4.cell(row=3, column=6).value = "Mult %"
sheet4.cell(row=3, column=7).value = "Me Win"
sheet4.cell(row=3, column=8).value = "Me Loss"
sheet4.cell(row=3, column=9).value = "Me %"
sheet4.cell(row=3, column=10).value = "Odds Win"
sheet4.cell(row=3, column=11).value = "Odds Loss"
sheet4.cell(row=3, column=12).value = "Odds %"
sheet4.cell(row=3, column=13).value = "Time"


# Write the values to the next row
sheet4.cell(row=next_row, column=1).value = Record['Win Reg'][0]
sheet4.cell(row=next_row, column=2).value = Record['Loss Reg'][0]
sheet4.cell(row=next_row, column=3).value = Record['Reg %'][0]
sheet4.cell(row=next_row, column=4).value = Record['Win Mult'][0]
sheet4.cell(row=next_row, column=5).value = Record['Loss Mult'][0]
sheet4.cell(row=next_row, column=6).value = Record['Mult %'][0]
sheet4.cell(row=next_row, column=7).value = Record['Me Win'][0]
sheet4.cell(row=next_row, column=8).value = Record['Me Loss'][0]
sheet4.cell(row=next_row, column=9).value = Record['Me %'][0]
sheet4.cell(row=next_row, column=10).value = Record['Odds Win'][0]
sheet4.cell(row=next_row, column=11).value = Record['Odds Loss'][0]
sheet4.cell(row=next_row, column=12).value = Record['Odds %'][0]
sheet4.cell(row=next_row, column=13).value = now

workbook.save('outputexcel.xlsx')












