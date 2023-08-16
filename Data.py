import requests
from bs4 import BeautifulSoup
import pandas as pd
from IPython.display import display
import os
import re
import math
from sklearn.preprocessing import MinMaxScaler
import csv
import openpyxl
from openpyxl import Workbook

scaler = MinMaxScaler()

pd.set_option('display.max_rows', None) # So you can view the whole table
pd.set_option('display.max_columns', None) # So you can view the whole table

url_list1 = [
    "https://www.teamrankings.com/ncaa-basketball/stat/points-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/average-scoring-margin",
    "https://www.teamrankings.com/ncaa-basketball/stat/offensive-efficiency",
    "https://www.teamrankings.com/ncaa-basketball/stat/floor-percentage",
    "https://www.teamrankings.com/ncaa-basketball/stat/1st-half-points-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/2nd-half-points-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/overtime-points-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/average-1st-half-margin",
    "https://www.teamrankings.com/ncaa-basketball/stat/average-2nd-half-margin",
    "https://www.teamrankings.com/ncaa-basketball/stat/average-overtime-margin",
    "https://www.teamrankings.com/ncaa-basketball/stat/points-from-2-pointers",
    "https://www.teamrankings.com/ncaa-basketball/stat/points-from-3-pointers",
    "https://www.teamrankings.com/ncaa-basketball/stat/percent-of-points-from-2-pointers",
    "https://www.teamrankings.com/ncaa-basketball/stat/percent-of-points-from-3-pointers",
    "https://www.teamrankings.com/ncaa-basketball/stat/percent-of-points-from-free-throws",
    "https://www.teamrankings.com/ncaa-basketball/stat/shooting-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/effective-field-goal-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/three-point-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/two-point-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/free-throw-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/true-shooting-percentage",
    "https://www.teamrankings.com/ncaa-basketball/stat/field-goals-made-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/field-goals-attempted-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/three-pointers-made-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/three-pointers-attempted-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/free-throws-made-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/free-throws-attempted-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/three-point-rate",
    "https://www.teamrankings.com/ncaa-basketball/stat/two-point-rate",
    "https://www.teamrankings.com/ncaa-basketball/stat/fta-per-fga",
    "https://www.teamrankings.com/ncaa-basketball/stat/ftm-per-100-possessions",
    "https://www.teamrankings.com/ncaa-basketball/stat/free-throw-rate",
    "https://www.teamrankings.com/ncaa-basketball/stat/non-blocked-2-pt-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/offensive-rebounds-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/defensive-rebounds-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/team-rebounds-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/total-rebounds-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/offensive-rebounding-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/defensive-rebounding-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/total-rebounding-percentage",
    "https://www.teamrankings.com/ncaa-basketball/stat/blocks-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/steals-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/block-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/steals-perpossession",
    "https://www.teamrankings.com/ncaa-basketball/stat/steal-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/assists-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/turnovers-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/turnovers-per-possession",
    "https://www.teamrankings.com/ncaa-basketball/stat/assist--per--turnover-ratio",
    "https://www.teamrankings.com/ncaa-basketball/stat/assists-per-fgm",
    "https://www.teamrankings.com/ncaa-basketball/stat/assists-per-possession",
    "https://www.teamrankings.com/ncaa-basketball/stat/turnover-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/personal-fouls-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/personal-fouls-per-possession",
    "https://www.teamrankings.com/ncaa-basketball/stat/personal-foul-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-points-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-average-scoring-margin",
    "https://www.teamrankings.com/ncaa-basketball/stat/defensive-efficiency",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-floor-percentage",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-1st-half-points-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-2nd-half-points-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-overtime-points-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-points-from-2-pointers",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-points-from-3-pointers",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-percent-of-points-from-2-pointers",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-percent-of-points-from-3-pointers",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-percent-of-points-from-free-throws",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-shooting-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-effective-field-goal-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-three-point-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-two-point-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-free-throw-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-true-shooting-percentage",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-field-goals-made-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-field-goals-attempted-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-three-pointers-made-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-three-pointers-attempted-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-free-throws-made-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-free-throws-attempted-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-three-point-rate",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-two-point-rate",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-fta-per-fga",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-ftm-per-100-possessions",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-free-throw-rate",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-non-blocked-2-pt-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-offensive-rebounds-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-defensive-rebounds-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-team-rebounds-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-total-rebounds-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-offensive-rebounding-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-defensive-rebounding-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-blocks-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-steals-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-block-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-steals-perpossession",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-steal-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-assists-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-turnovers-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-assist--per--turnover-ratio",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-assists-per-fgm",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-assists-per-possession",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-turnovers-per-possession",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-turnover-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-personal-fouls-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-personal-fouls-per-possession",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-personal-foul-pct",
    "https://www.teamrankings.com/ncaa-basketball/stat/possessions-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/extra-chances-per-game",
    "https://www.teamrankings.com/ncaa-basketball/stat/effective-possession-ratio",
    "https://www.teamrankings.com/ncaa-basketball/stat/opponent-effective-possession-ratio", #1
]

url_list2 = [
    "https://www.teamrankings.com/ncaa-basketball/ranking/predictive-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/home-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/away-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/neutral-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/home-adv-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/luck-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/consistency-by-other",
]

url_list3 = [
    "https://www.teamrankings.com/ncaa-basketball/ranking/schedule-strength-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/future-sos-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/season-sos-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/sos-basic-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/in-conference-sos-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/non-conference-sos-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/last-5-games-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/last-10-games-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/in-conference-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/non-conference-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/vs-1-25-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/vs-26-50-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/vs-51-100-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/vs-101-200-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/vs-201-and-up-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/first-half-by-other",
    "https://www.teamrankings.com/ncaa-basketball/ranking/second-half-by-other",
]
# List of URL strings to scrape

Schedule = "https://www.teamrankings.com/ncb/schedules/?date=2023-03-12"
# For loop used to go through each of the url's in the list

WebContentsSchedule = requests.get(Schedule)
# Gets the content from the website

ReadContentsSchedule = BeautifulSoup(WebContentsSchedule.content, "html5lib")
# Uses Beautiful Soup to parse through the url contents

FindTableSchedule = ReadContentsSchedule.find("table")
# Uses Beautiful Soup to go through the website content and find "table"
# Do something with the data here, such as storing it in a list or a DataFrame

PresentTableSchedule = pd.read_html(str(FindTableSchedule), flavor="html5lib")[0]
# Uses Pandas to read the "FindTable" data and takes the information to present a table found in the HTML
# html5lib is a library used to parse the data??? and the [0] is used to access the first element of the table

Mask = PresentTableSchedule['Hotness Score'] != '--'
PresentTableSchedule = PresentTableSchedule.drop(PresentTableSchedule[~Mask].index)
regex = r"#\d+"
PresentTableSchedule['Matchup'] = PresentTableSchedule['Matchup'].apply(lambda x: re.sub(regex, '', x))
PresentTableSchedule['Matchup'] = PresentTableSchedule['Matchup'].str.strip()
Teams = PresentTableSchedule['Matchup'].str.split(re.compile(r'\s+at\s+|\s+vs\.\s+'))
#Teams.drop(Teams.index[:4], inplace=True)

for row in Teams:
    #Team1 = row[0]
    #Team2 = row[1]
    Team1 = "Purdue"
    Team2 = "F Dickinson"
    Team1 = Team1.strip()
    Team2 = Team2.strip()


    def calculate_win_percentage(x):
        try:
            # Split the value on the hyphen character "-"
            x = x.split("-")

            # Convert the number of wins and losses to integers
            x = [int(x[0]), int(x[1])]

            # Calculate the win percentage
            win_percentage = x[0] / (x[0] + x[1])

            # Round the win percentage to two decimal places
            win_percentage = round(win_percentage, 2)

            return win_percentage

        except (ValueError, ZeroDivisionError):
            # Return 0 if there is an error
            return 0


    Team1Table1Norm = pd.DataFrame()  # Empty Data Frame to be filled with rows
    Team1Table2Norm = pd.DataFrame()  # Empty Data Frame to be filled with rows
    Team1Table3Norm = pd.DataFrame()  # Empty Data Frame to be filled with rows
    Team2Table4Norm = pd.DataFrame()  # Empty Data Frame to be filled with rows
    Team2Table5Norm = pd.DataFrame()  # Empty Data Frame to be filled with rows
    Team2Table6Norm = pd.DataFrame()  # Empty Data Frame to be filled with rows

    for url in url_list1:
        # For loop used to go through each of the url's in the list

        WebContentsNorm = requests.get(url)
        # Gets the content from the website

        ReadContentsNorm = BeautifulSoup(WebContentsNorm.content, "html5lib")
        # Uses Beautiful Soup to parse through the url contents

        FindTableNorm = ReadContentsNorm.find("table")
        # Uses Beautiful Soup to go through the website content and find "table"
        # Do something with the data here, such as storing it in a list or a DataFrame

        PresentTableNorm = pd.read_html(str(FindTableNorm), flavor="html5lib")[0]
        # Uses Pandas to read the "FindTable" data and takes the information to present a table found in the HTML
        # html5lib is a library used to parse the data??? and the [0] is used to access the first element of the table
        PresentTableNorm['2022'] = PresentTableNorm['2022'].apply(lambda x: str(x).replace('%', ''))
        PresentTableNorm['Last 3'] = PresentTableNorm['Last 3'].apply(lambda x: str(x).replace('%', ''))
        PresentTableNorm['Last 1'] = PresentTableNorm['Last 1'].apply(lambda x: str(x).replace('%', ''))
        PresentTableNorm['Home'] = PresentTableNorm['Home'].apply(lambda x: str(x).replace('%', ''))
        PresentTableNorm['Away'] = PresentTableNorm['Away'].apply(lambda x: str(x).replace('%', ''))

        PresentTableNorm = PresentTableNorm.replace("--", 0, regex=True)

        PresentTableNorm['2022'] = scaler.fit_transform(PresentTableNorm[['2022']])
        PresentTableNorm['Last 3'] = scaler.fit_transform(PresentTableNorm[['Last 3']])
        PresentTableNorm['Last 1'] = scaler.fit_transform(PresentTableNorm[['Last 1']])
        PresentTableNorm['Home'] = scaler.fit_transform(PresentTableNorm[['Home']])
        PresentTableNorm['Away'] = scaler.fit_transform(PresentTableNorm[['Away']])

        rowNorm = PresentTableNorm.loc[PresentTableNorm['Team'] == Team1]
        # Locates the row where the Team1 variable is equal to one of the names in the table, and then extracts the information from that row and stores it in the variable

        rowNorm = rowNorm.assign(stat=os.path.basename(url))
        # Adds a new column to the row named "stat" and it takes the end of the url being used and adds it to the "stat" column

        Team1Table1Norm = pd.concat([Team1Table1Norm, rowNorm], ignore_index=True)
        # Team1Table is an empty dataframe and each time the loop passes through, Team1Table equals Team1Table + the row that was extracted in that pass of the loop from its specific url
        # Each pass of the loop adds a row to Team1Table

        Team1Table1Norm = Team1Table1Norm.drop(columns=["2021"])
        # Removes the column of stats from last year

    for url in url_list2:
        # For loop used to go through each of the url's in the list

        WebContents2Norm = requests.get(url)
        # Gets the content from the website

        ReadContents2Norm = BeautifulSoup(WebContents2Norm.content, "html5lib")
        # Uses Beautiful Soup to parse through the url contents

        FindTable2Norm = ReadContents2Norm.find("table")
        # Uses Beautiful Soup to go through the website content and find "table"
        # Do something with the data here, such as storing it in a list or a DataFrame

        PresentTable2Norm = pd.read_html(str(FindTable2Norm), flavor="html5lib")[0]
        # Uses Pandas to read the "FindTable" data and takes the information to present a table found in the HTML
        # html5lib is a library used to parse the data??? and the [0] is used to access the first element of the table

        PresentTable2Norm['Rating'] = PresentTable2Norm['Rating'].apply(lambda x: str(x).replace('%', ''))
        PresentTable2Norm['v 1-25'] = PresentTable2Norm['v 1-25'].apply(lambda x: str(x).replace('%', ''))
        PresentTable2Norm['v 26-50'] = PresentTable2Norm['v 26-50'].apply(lambda x: str(x).replace('%', ''))
        PresentTable2Norm['v 51-100'] = PresentTable2Norm['v 51-100'].apply(lambda x: str(x).replace('%', ''))
        PresentTable2Norm['Hi'] = PresentTable2Norm['Hi'].apply(lambda x: str(x).replace('%', ''))
        PresentTable2Norm['Low'] = PresentTable2Norm['Low'].apply(lambda x: str(x).replace('%', ''))
        PresentTable2Norm['Last'] = PresentTable2Norm['Last'].apply(lambda x: str(x).replace('%', ''))

        PresentTable2Norm['v 1-25'] = PresentTable2Norm['v 1-25'].apply(calculate_win_percentage)
        PresentTable2Norm['v 26-50'] = PresentTable2Norm['v 26-50'].apply(calculate_win_percentage)
        PresentTable2Norm['v 51-100'] = PresentTable2Norm['v 51-100'].apply(calculate_win_percentage)

        PresentTable2Norm = PresentTable2Norm.replace("--", 0, regex=True)

        PresentTable2Norm['Rating'] = scaler.fit_transform(PresentTable2Norm[['Rating']])
        PresentTable2Norm['v 1-25'] = scaler.fit_transform(PresentTable2Norm[['v 1-25']])
        PresentTable2Norm['v 26-50'] = scaler.fit_transform(PresentTable2Norm[['v 26-50']])
        PresentTable2Norm['v 51-100'] = scaler.fit_transform(PresentTable2Norm[['v 51-100']])
        PresentTable2Norm['Hi'] = scaler.fit_transform(PresentTable2Norm[['Hi']])
        PresentTable2Norm['Low'] = scaler.fit_transform(PresentTable2Norm[['Low']])
        PresentTable2Norm['Last'] = scaler.fit_transform(PresentTable2Norm[['Last']])

        # RemoveComponent1 = lambda x: x.split(" (")[0]
        # PresentTable2["Team"] = PresentTable2["Team"].apply(RemoveComponent1)
        PresentTable2Norm['Team'] = PresentTable2Norm['Team'].str.replace(r'\((?!FL|OH|NY|PA\)).*\)', '', regex=True)
        # Removes the record from the team column so the team can be found when indexing/matching
        # Also keeps the parentheses if FL OH NY or PA is inside of them

        PresentTable2Norm['Team'] = PresentTable2Norm['Team'].str.strip()
        # Removes the empty space in the team names after their records were removed from their name in the team column
        # This allows the next line of code to work and the team to be properly matched

        row2Norm = PresentTable2Norm.loc[PresentTable2Norm['Team'] == Team1]
        # Locates the row where the Team1 variable is equal to one of the names in the table, and then extracts the information from that row and stores it in the variable

        row2Norm = row2Norm.assign(stat=os.path.basename(url))
        # Adds a new column to the row named "stat" and it takes the end of the url being used and adds it to the "stat" column

        Team1Table2Norm = pd.concat([Team1Table2Norm, row2Norm], ignore_index=True)
        # Team1Table is an empty dataframe and each time the loop passes through, Team1Table equals Team1Table + the row that was extracted in that pass of the loop from its specific url
        # Each pass of the loop adds a row to Team1Table

        # Team1Table2 = Team1Table2.drop(columns=[""])
        # Removes the column of stats from last year

    for url in url_list3:
        # For loop used to go through each of the url's in the list

        WebContents3Norm = requests.get(url)
        # Gets the content from the website

        ReadContents3Norm = BeautifulSoup(WebContents3Norm.content, "html5lib")
        # Uses Beautiful Soup to parse through the url contents

        FindTable3Norm = ReadContents3Norm.find("table")
        # Uses Beautiful Soup to go through the website content and find "table"
        # Do something with the data here, such as storing it in a list or a DataFrame

        PresentTable3Norm = pd.read_html(str(FindTable3Norm), flavor="html5lib")[0]
        # Uses Pandas to read the "FindTable" data and takes the information to present a table found in the HTML
        # html5lib is a library used to parse the data??? and the [0] is used to access the first element of the table

        PresentTable3Norm['Rating'] = PresentTable3Norm['Rating'].apply(lambda x: str(x).replace('%', ''))
        PresentTable3Norm['Hi'] = PresentTable3Norm['Hi'].apply(lambda x: str(x).replace('%', ''))
        PresentTable3Norm['Low'] = PresentTable3Norm['Low'].apply(lambda x: str(x).replace('%', ''))
        PresentTable3Norm['Last'] = PresentTable3Norm['Last'].apply(lambda x: str(x).replace('%', ''))

        PresentTable3Norm = PresentTable3Norm.replace("--", 0, regex=True)

        PresentTable3Norm['Rating'] = scaler.fit_transform(PresentTable3Norm[['Rating']])
        PresentTable3Norm['Hi'] = scaler.fit_transform(PresentTable3Norm[['Hi']])
        PresentTable3Norm['Low'] = scaler.fit_transform(PresentTable3Norm[['Low']])
        PresentTable3Norm['Last'] = scaler.fit_transform(PresentTable3Norm[['Last']])

        # RemoveComponent2 = lambda x: x.split(" (")[0]
        PresentTable3Norm['Team'] = PresentTable3Norm['Team'].str.replace(r'\((?!FL|OH|NY|PA\)).*\)', '', regex=True)
        # Removes the record from the team column so the team can be found when indexing/matching
        # Also keeps the parentheses if FL OH NY or PA is inside of them

        PresentTable3Norm['Team'] = PresentTable3Norm['Team'].str.strip()
        # Removes the empty space in the team names after their records were removed from their name in the team column
        # This allows the next line of code to work and the team to be properly matched

        row3Norm = PresentTable3Norm.loc[PresentTable3Norm['Team'] == Team1]
        # Locates the row where the Team1 variable is equal to one of the names in the table, and then extracts the information from that row and stores it in the variable

        row3Norm = row3Norm.assign(stat=os.path.basename(url))
        # Adds a new column to the row named "stat" and it takes the end of the url being used and adds it to the "stat" column

        Team1Table3Norm = pd.concat([Team1Table3Norm, row3Norm], ignore_index=True)
        # Team1Table is an empty dataframe and each time the loop passes through, Team1Table equals Team1Table + the row that was extracted in that pass of the loop from its specific url
        # Each pass of the loop adds a row to Team1Table

        # Team1Table3 = Team1Table3.drop(columns=[""])
        # Removes the column of stats from last year

    for url in url_list1:
        # For loop used to go through each of the url's in the list

        WebContents4Norm = requests.get(url)
        # Gets the content from the website

        ReadContents4Norm = BeautifulSoup(WebContents4Norm.content, "html5lib")
        # Uses Beautiful Soup to parse through the url contents

        FindTable4Norm = ReadContents4Norm.find("table")
        # Uses Beautiful Soup to go through the website content and find "table"
        # Do something with the data here, such as storing it in a list or a DataFrame

        PresentTable4Norm = pd.read_html(str(FindTable4Norm), flavor="html5lib")[0]
        # Uses Pandas to read the "FindTable" data and takes the information to present a table found in the HTML
        # html5lib is a library used to parse the data??? and the [0] is used to access the first element of the table

        PresentTable4Norm['2022'] = PresentTable4Norm['2022'].apply(lambda x: str(x).replace('%', ''))
        PresentTable4Norm['Last 3'] = PresentTable4Norm['Last 3'].apply(lambda x: str(x).replace('%', ''))
        PresentTable4Norm['Last 1'] = PresentTable4Norm['Last 1'].apply(lambda x: str(x).replace('%', ''))
        PresentTable4Norm['Home'] = PresentTable4Norm['Home'].apply(lambda x: str(x).replace('%', ''))
        PresentTable4Norm['Away'] = PresentTable4Norm['Away'].apply(lambda x: str(x).replace('%', ''))

        PresentTable4Norm = PresentTable4Norm.replace("--", 0, regex=True)

        PresentTable4Norm['2022'] = scaler.fit_transform(PresentTable4Norm[['2022']])
        PresentTable4Norm['Last 3'] = scaler.fit_transform(PresentTable4Norm[['Last 3']])
        PresentTable4Norm['Last 1'] = scaler.fit_transform(PresentTable4Norm[['Last 1']])
        PresentTable4Norm['Home'] = scaler.fit_transform(PresentTable4Norm[['Home']])
        PresentTable4Norm['Away'] = scaler.fit_transform(PresentTable4Norm[['Away']])

        row4Norm = PresentTable4Norm.loc[PresentTable4Norm['Team'] == Team2]
        # Locates the row where the Team1 variable is equal to one of the names in the table, and then extracts the information from that row and stores it in the variable

        row4Norm = row4Norm.assign(stat=os.path.basename(url))
        # Adds a new column to the row named "stat" and it takes the end of the url being used and adds it to the "stat" column

        Team2Table4Norm = pd.concat([Team2Table4Norm, row4Norm], ignore_index=True)
        # Team1Table is an empty dataframe and each time the loop passes through, Team1Table equals Team1Table + the row that was extracted in that pass of the loop from its specific url
        # Each pass of the loop adds a row to Team1Table

        Team2Table4Norm = Team2Table4Norm.drop(columns=["2021"])
        # Removes the column of stats from last year

    for url in url_list2:
        # For loop used to go through each of the url's in the list

        WebContents5Norm = requests.get(url)
        # Gets the content from the website

        ReadContents5Norm = BeautifulSoup(WebContents5Norm.content, "html5lib")
        # Uses Beautiful Soup to parse through the url contents

        FindTable5Norm = ReadContents5Norm.find("table")
        # Uses Beautiful Soup to go through the website content and find "table"
        # Do something with the data here, such as storing it in a list or a DataFrame

        PresentTable5Norm = pd.read_html(str(FindTable5Norm), flavor="html5lib")[0]
        # Uses Pandas to read the "FindTable" data and takes the information to present a table found in the HTML
        # html5lib is a library used to parse the data??? and the [0] is used to access the first element of the table

        PresentTable5Norm['Rating'] = PresentTable5Norm['Rating'].apply(lambda x: str(x).replace('%', ''))
        PresentTable5Norm['v 1-25'] = PresentTable5Norm['v 1-25'].apply(lambda x: str(x).replace('%', ''))
        PresentTable5Norm['v 26-50'] = PresentTable5Norm['v 26-50'].apply(lambda x: str(x).replace('%', ''))
        PresentTable5Norm['v 51-100'] = PresentTable5Norm['v 51-100'].apply(lambda x: str(x).replace('%', ''))
        PresentTable5Norm['Hi'] = PresentTable5Norm['Hi'].apply(lambda x: str(x).replace('%', ''))
        PresentTable5Norm['Low'] = PresentTable5Norm['Low'].apply(lambda x: str(x).replace('%', ''))
        PresentTable5Norm['Last'] = PresentTable5Norm['Last'].apply(lambda x: str(x).replace('%', ''))

        PresentTable5Norm['v 1-25'] = PresentTable5Norm['v 1-25'].apply(calculate_win_percentage)
        PresentTable5Norm['v 26-50'] = PresentTable5Norm['v 26-50'].apply(calculate_win_percentage)
        PresentTable5Norm['v 51-100'] = PresentTable5Norm['v 51-100'].apply(calculate_win_percentage)

        PresentTable5Norm = PresentTable5Norm.replace("--", 0, regex=True)

        PresentTable5Norm['Rating'] = scaler.fit_transform(PresentTable5Norm[['Rating']])
        PresentTable5Norm['v 1-25'] = scaler.fit_transform(PresentTable5Norm[['v 1-25']])
        PresentTable5Norm['v 26-50'] = scaler.fit_transform(PresentTable5Norm[['v 26-50']])
        PresentTable5Norm['v 51-100'] = scaler.fit_transform(PresentTable5Norm[['v 51-100']])
        PresentTable5Norm['Hi'] = scaler.fit_transform(PresentTable5Norm[['Hi']])
        PresentTable5Norm['Low'] = scaler.fit_transform(PresentTable5Norm[['Low']])
        PresentTable5Norm['Last'] = scaler.fit_transform(PresentTable5Norm[['Last']])

        # RemoveComponent1 = lambda x: x.split(" (")[0]
        # PresentTable2["Team"] = PresentTable2["Team"].apply(RemoveComponent1)
        PresentTable5Norm['Team'] = PresentTable5Norm['Team'].str.replace(r'\((?!FL|OH|NY|PA\)).*\)', '', regex=True)
        # Removes the record from the team column so the team can be found when indexing/matching
        # Also keeps the parentheses if FL OH NY or PA is inside of them

        PresentTable5Norm['Team'] = PresentTable5Norm['Team'].str.strip()
        # Removes the empty space in the team names after their records were removed from their name in the team column
        # This allows the next line of code to work and the team to be properly matched

        row5Norm = PresentTable5Norm.loc[PresentTable5Norm['Team'] == Team2]
        # Locates the row where the Team1 variable is equal to one of the names in the table, and then extracts the information from that row and stores it in the variable

        row5Norm = row5Norm.assign(stat=os.path.basename(url))
        # Adds a new column to the row named "stat" and it takes the end of the url being used and adds it to the "stat" column

        Team2Table5Norm = pd.concat([Team2Table5Norm, row5Norm], ignore_index=True)
        # Team1Table is an empty dataframe and each time the loop passes through, Team1Table equals Team1Table + the row that was extracted in that pass of the loop from its specific url
        # Each pass of the loop adds a row to Team1Table

        # Team1Table2 = Team1Table2.drop(columns=[""])
        # Removes the column of stats from last year

    for url in url_list3:
        # For loop used to go through each of the url's in the list

        WebContents6Norm = requests.get(url)
        # Gets the content from the website

        ReadContents6Norm = BeautifulSoup(WebContents6Norm.content, "html5lib")
        # Uses Beautiful Soup to parse through the url contents

        FindTable6Norm = ReadContents6Norm.find("table")
        # Uses Beautiful Soup to go through the website content and find "table"
        # Do something with the data here, such as storing it in a list or a DataFrame

        PresentTable6Norm = pd.read_html(str(FindTable6Norm), flavor="html5lib")[0]
        # Uses Pandas to read the "FindTable" data and takes the information to present a table found in the HTML
        # html5lib is a library used to parse the data??? and the [0] is used to access the first element of the table

        PresentTable6Norm['Rating'] = PresentTable6Norm['Rating'].apply(lambda x: str(x).replace('%', ''))
        PresentTable6Norm['Hi'] = PresentTable6Norm['Hi'].apply(lambda x: str(x).replace('%', ''))
        PresentTable6Norm['Low'] = PresentTable6Norm['Low'].apply(lambda x: str(x).replace('%', ''))
        PresentTable6Norm['Last'] = PresentTable6Norm['Last'].apply(lambda x: str(x).replace('%', ''))

        PresentTable6Norm = PresentTable6Norm.replace("--", 0, regex=True)

        PresentTable6Norm['Rating'] = scaler.fit_transform(PresentTable6Norm[['Rating']])
        PresentTable6Norm['Hi'] = scaler.fit_transform(PresentTable6Norm[['Hi']])
        PresentTable6Norm['Low'] = scaler.fit_transform(PresentTable6Norm[['Low']])
        PresentTable6Norm['Last'] = scaler.fit_transform(PresentTable6Norm[['Last']])

        # RemoveComponent2 = lambda x: x.split(" (")[0]
        PresentTable6Norm['Team'] = PresentTable6Norm['Team'].str.replace(r'\((?!FL|OH|NY|PA\)).*\)', '', regex=True)
        # Removes the record from the team column so the team can be found when indexing/matching
        # Also keeps the parentheses if FL OH NY or PA is inside of them

        PresentTable6Norm['Team'] = PresentTable6Norm['Team'].str.strip()
        # Removes the empty space in the team names after their records were removed from their name in the team column
        # This allows the next line of code to work and the team to be properly matched

        row6Norm = PresentTable6Norm.loc[PresentTable6Norm['Team'] == Team2]
        # Locates the row where the Team1 variable is equal to one of the names in the table, and then extracts the information from that row and stores it in the variable

        row6Norm = row6Norm.assign(stat=os.path.basename(url))
        # Adds a new column to the row named "stat" and it takes the end of the url being used and adds it to the "stat" column

        Team2Table6Norm = pd.concat([Team2Table6Norm, row6Norm], ignore_index=True)
        # Team1Table is an empty dataframe and each time the loop passes through, Team1Table equals Team1Table + the row that was extracted in that pass of the loop from its specific url
        # Each pass of the loop adds a row to Team1Table

        # Team1Table3 = Team1Table3.drop(columns=[""])
        # Removes the column of stats from last year

    Team1Table1Norm['2022'] = pd.to_numeric(Team1Table1Norm['2022'])
    Team1Table1Norm['Last 3'] = pd.to_numeric(Team1Table1Norm['Last 3'])
    Team1Table1Norm['Last 1'] = pd.to_numeric(Team1Table1Norm['Last 1'])
    Team1Table1Norm['Home'] = pd.to_numeric(Team1Table1Norm['Home'])
    Team1Table1Norm['Away'] = pd.to_numeric(Team1Table1Norm['Away'])

    Team1Table2Norm['Rating'] = pd.to_numeric(Team1Table2Norm['Rating'], errors='ignore')
    Team1Table2Norm['v 1-25'] = pd.to_numeric(Team1Table2Norm['v 1-25'], errors='ignore')
    Team1Table2Norm['v 26-50'] = pd.to_numeric(Team1Table2Norm['v 26-50'], errors='ignore')
    Team1Table2Norm['v 51-100'] = pd.to_numeric(Team1Table2Norm['v 51-100'], errors='ignore')
    Team1Table2Norm['Hi'] = pd.to_numeric(Team1Table2Norm['Hi'], errors='ignore')
    Team1Table2Norm['Low'] = pd.to_numeric(Team1Table2Norm['Low'], errors='ignore')
    Team1Table2Norm['Last'] = pd.to_numeric(Team1Table2Norm['Last'], errors='ignore')

    Team1Table3Norm['Rating'] = pd.to_numeric(Team1Table3Norm['Rating'], errors='ignore')
    Team1Table3Norm['Hi'] = pd.to_numeric(Team1Table3Norm['Hi'], errors='ignore')
    Team1Table3Norm['Low'] = pd.to_numeric(Team1Table3Norm['Low'], errors='ignore')
    Team1Table3Norm['Last'] = pd.to_numeric(Team1Table3Norm['Last'], errors='ignore')

    Team2Table4Norm['2022'] = pd.to_numeric(Team2Table4Norm['2022'], errors='ignore')
    Team2Table4Norm['Last 3'] = pd.to_numeric(Team2Table4Norm['Last 3'], errors='ignore')
    Team2Table4Norm['Last 1'] = pd.to_numeric(Team2Table4Norm['Last 1'], errors='ignore')
    Team2Table4Norm['Home'] = pd.to_numeric(Team2Table4Norm['Home'], errors='ignore')
    Team2Table4Norm['Away'] = pd.to_numeric(Team2Table4Norm['Away'], errors='ignore')

    Team2Table5Norm['Rating'] = pd.to_numeric(Team2Table5Norm['Rating'], errors='ignore')
    Team2Table5Norm['v 1-25'] = pd.to_numeric(Team2Table5Norm['v 1-25'], errors='ignore')
    Team2Table5Norm['v 26-50'] = pd.to_numeric(Team2Table5Norm['v 26-50'], errors='ignore')
    Team2Table5Norm['v 51-100'] = pd.to_numeric(Team2Table5Norm['v 51-100'], errors='ignore')
    Team2Table5Norm['Hi'] = pd.to_numeric(Team2Table5Norm['Hi'], errors='ignore')
    Team2Table5Norm['Low'] = pd.to_numeric(Team2Table5Norm['Low'], errors='ignore')
    Team2Table5Norm['Last'] = pd.to_numeric(Team2Table5Norm['Last'], errors='ignore')

    Team2Table6Norm['Rating'] = pd.to_numeric(Team2Table6Norm['Rating'], errors='ignore')
    Team2Table6Norm['Hi'] = pd.to_numeric(Team2Table6Norm['Hi'], errors='ignore')
    Team2Table6Norm['Low'] = pd.to_numeric(Team2Table6Norm['Low'], errors='ignore')
    Team2Table6Norm['Last'] = pd.to_numeric(Team2Table6Norm['Last'], errors='ignore')
    # Changes all of the values in the table to a numerical value instead of a string so it can be compared

    Compare1Norm = Team1Table1Norm[['2022', 'Last 3', 'Last 1', 'Home', 'Away']].subtract(
        Team2Table4Norm[['2022', 'Last 3', 'Last 1', 'Home', 'Away']])
    Compare2Norm = Team1Table2Norm[['Rating', 'v 1-25', 'v 26-50', 'v 51-100', 'Hi', 'Low', 'Last']].subtract(
        Team2Table5Norm[['Rating', 'v 1-25', 'v 26-50', 'v 51-100', 'Hi', 'Low', 'Last']], fill_value=0)
    Compare3Norm = Team1Table3Norm[['Rating', 'Hi', 'Low', 'Last']].subtract(
        Team2Table6Norm[['Rating', 'Hi', 'Low', 'Last']], fill_value=0)
    # Subtracts all the corresponding values in both tables from each other and stores that table in a new variable called Compare1

    row46to47Norm = Compare1Norm.iloc[46:47, :]
    row46to47Norm = row46to47Norm * -1
    Compare1Norm.iloc[46:47, :] = row46to47Norm
    row51to96Norm = Compare1Norm.iloc[51:96, :]
    row51to96Norm = row51to96Norm * -1
    Compare1Norm.iloc[51:96, :] = row51to96Norm
    row98to100Norm = Compare1Norm.iloc[98:100, :]
    row98to100Norm = row98to100Norm * -1
    Compare1Norm.iloc[98:100, :] = row98to100Norm
    row109Norm = Compare1Norm.iloc[109, :]
    row109Norm = row109Norm * -1
    Compare1Norm.iloc[109, :] = row109Norm
    # Multiplies certain rows by -1 because those rows should be added subtracted the opposite way to make sense

    positives2022Norm = 0
    negatives2022Norm = 0
    positivesLast3Norm = 0
    negativesLast3Norm = 0
    positivesLast1Norm = 0
    negativesLast1Norm = 0
    positivesHomeNorm = 0
    negativesHomeNorm = 0
    positivesAwayNorm = 0
    negativesAwayNorm = 0
    x2022Norm = 0
    xLast3Norm = 0
    xLast1Norm = 0
    xHomeNorm = 0
    xAwayNorm = 0

    positivesRating1Norm = 0
    positives1to25Norm = 0
    positives26to50Norm = 0
    positives51to100Norm = 0
    positives1ofHiNorm = 0
    positives1ofLowNorm = 0
    positives1ofLastNorm = 0
    negativesRating1Norm = 0
    negatives1to25Norm = 0
    negatives26to50Norm = 0
    negatives51to100Norm = 0
    negatives1ofHiNorm = 0
    negatives1ofLowNorm = 0
    negatives1ofLastNorm = 0
    xRating1Norm = 0
    xof1to25Norm = 0
    xof26to50Norm = 0
    xof51to100Norm = 0
    x1ofHiNorm = 0
    x1ofLowNorm = 0
    x1ofLastNorm = 0

    positivesRating2Norm = 0
    positives2ofHiNorm = 0
    positives2ofLowNorm = 0
    positives2ofLastNorm = 0
    negativesRating2Norm = 0
    negatives2ofHiNorm = 0
    negatives2ofLowNorm = 0
    negatives2ofLastNorm = 0
    xRating2Norm = 0
    x2ofHiNorm = 0
    x2ofLowNorm = 0
    x2ofLastNorm = 0

    #StatRecords1Norm = pd.DataFrame()  # Empty DataFrame to be filled with rows
    #StatRecords2Norm = pd.DataFrame()  # Empty DataFrame to be filled with rows
    #StatRecords3Norm = pd.DataFrame()  # Empty DataFrame to be filled with rows
    StatRecords1Norm = pd.read_csv("SR1Norm.csv")
    StatRecords2Norm = pd.read_csv("SR2Norm.csv")
    StatRecords3Norm = pd.read_csv("SR3Norm.csv")

    #StatNameNorm = os.path.basename(url)
        #StatRecords1Norm = pd.concat([StatRecords1Norm, pd.DataFrame({"Stat Name": [StatNameNorm]})], ignore_index=True)
        #StatRecords1Norm = StatRecords1Norm.assign(
            #**{"2022 Win": 0, "2022 Loss": 0, "2022 Win %": 0, "Last 3 Win": 0, "Last 3 Loss": 0, "Last 3 Win %": 0,
               #"Last 1 Win": 0, "Last 1 Loss": 0, "Last 1 Win %": 0, "Home Win": 0, "Home Loss": 0, "Home Win %": 0,
               #"Away Win": 0, "Away Loss": 0, "Away Win %": 0})

    #for url in url_list2:
        #StatNameNorm = os.path.basename(url)
        #StatRecords2Norm = pd.concat([StatRecords2Norm, pd.DataFrame({"Stat Name": [StatNameNorm]})], ignore_index=True)
        #StatRecords2Norm = StatRecords2Norm.assign(
            #**{"Rating Win": 0, "Rating Loss": 0, "Rating Win %": 0, "v 1-25 Win": 0, "v 1-25 Loss": 0,
               #"v 1-25 Win %": 0, "v 26-50 Win": 0, "v 26-50 Loss": 0, "v 26-50 Win %": 0, "v 51-100 Win": 0,
               #"v 51-100 Loss": 0, "v 51-100 Win %": 0, "Hi Win": 0, "Hi Loss": 0, "Hi Win %": 0, "Low Win": 0,
               #"Low Loss": 0, "Low Win %": 0, "Last Win": 0, "Last Loss": 0, "Last Win %": 0})

    #for url in url_list3:
        #StatNameNorm = os.path.basename(url)
        #StatRecords3Norm = pd.concat([StatRecords3Norm, pd.DataFrame({"Stat Name": [StatNameNorm]})], ignore_index=True)
        #StatRecords3Norm = StatRecords3Norm.assign(
            #**{"Rating Win": 0, "Rating Loss": 0, "Rating Win %": 0, "Hi Win": 0, "Hi Loss": 0, "Hi Win %": 0,
               #"Low Win": 0, "Low Loss": 0, "Low Win %": 0, "Last Win": 0, "Last Loss": 0, "Last Win %": 0})

    #GameOutcomeNorm = input("Enter a 1 for team 1 or a 2 for team 2: ")
    #GameOutcomeNorm = int(GameOutcomeNorm)
    #GameOutcomeNorm = 1


    workbook = openpyxl.load_workbook('outputexcel.xlsx')
    sheet = workbook.worksheets[0]

    # Iterate through the rows and columns in the sheet
    #for row in sheet.iter_rows():
        #for cell in row:
            # Check if the cell value is #NA and if the cell does not contain a formula
            #if cell.value == '#NA' and cell.data_type != 'f':
                # Set the cell value to 0
                #cell.value = 0

    # Save the workbook
    #workbook.save('outputexcel.xlsx')
    GameOutcomeNorm = 0

    for row in sheet:
        values = [cell.value for cell in row]
        try:
            if values[0] == Team1 and values[1] == Team2:
                GameOutcomeNorm = values[4]
                break
            else:
                GameOutcomeNorm = 0
        except IndexError:
            GameOutcomeNorm = 0

    GameOutcomeNorm = int(GameOutcomeNorm)

    while x2022Norm < len(Compare1Norm['2022']):
        try:
            # Access the value at the specified index
            value = Compare1Norm['2022'].iloc[x2022Norm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives2022Norm += 1
            if GameOutcomeNorm == 1:
                StatRecords1Norm['2022 Win'].iloc[x2022Norm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords1Norm['2022 Loss'].iloc[x2022Norm] += 1
        elif value < 0:
            negatives2022Norm += 1
            if GameOutcomeNorm == 2:
                StatRecords1Norm['2022 Win'].iloc[x2022Norm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords1Norm['2022 Loss'].iloc[x2022Norm] += 1

        if StatRecords1Norm['2022 Win'].iloc[x2022Norm] > StatRecords1Norm['2022 Loss'].iloc[x2022Norm]:
            StatRecords1Norm['2022 Win %'].iloc[x2022Norm] = (StatRecords1Norm['2022 Win'].iloc[x2022Norm]) / (
                        StatRecords1Norm['2022 Win'].iloc[x2022Norm] + StatRecords1Norm['2022 Loss'].iloc[x2022Norm])
        elif StatRecords1Norm['2022 Win'].iloc[x2022Norm] < StatRecords1Norm['2022 Loss'].iloc[x2022Norm]:
            StatRecords1Norm['2022 Win %'].iloc[x2022Norm] = ((StatRecords1Norm['2022 Loss'].iloc[x2022Norm]) / (
                        StatRecords1Norm['2022 Win'].iloc[x2022Norm] + StatRecords1Norm['2022 Loss'].iloc[
                    x2022Norm])) * -1
        elif StatRecords1Norm['2022 Win'].iloc[x2022Norm] == StatRecords1Norm['2022 Loss'].iloc[x2022Norm]:
            StatRecords1Norm['2022 Win %'].iloc[x2022Norm] = 0.5
        # Calculates win % of the stats
        x2022Norm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xLast3Norm < len(Compare1Norm['Last 3']):
        try:
            # Access the value at the specified index
            value = Compare1Norm['Last 3'].iloc[xLast3Norm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positivesLast3Norm += 1
            if GameOutcomeNorm == 1:
                StatRecords1Norm['Last 3 Win'].iloc[xLast3Norm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords1Norm['Last 3 Loss'].iloc[xLast3Norm] += 1
        elif value < 0:
            negativesLast3Norm += 1
            if GameOutcomeNorm == 2:
                StatRecords1Norm['Last 3 Win'].iloc[xLast3Norm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords1Norm['Last 3 Loss'].iloc[xLast3Norm] += 1

        if StatRecords1Norm['Last 3 Win'].iloc[xLast3Norm] > StatRecords1Norm['Last 3 Loss'].iloc[xLast3Norm]:
            StatRecords1Norm['Last 3 Win %'].iloc[xLast3Norm] = (StatRecords1Norm['Last 3 Win'].iloc[xLast3Norm]) / (
                        StatRecords1Norm['Last 3 Win'].iloc[xLast3Norm] + StatRecords1Norm['Last 3 Loss'].iloc[
                    xLast3Norm])
        elif StatRecords1Norm['Last 3 Win'].iloc[xLast3Norm] < StatRecords1Norm['Last 3 Loss'].iloc[xLast3Norm]:
            StatRecords1Norm['Last 3 Win %'].iloc[xLast3Norm] = ((StatRecords1Norm['Last 3 Loss'].iloc[xLast3Norm]) / (
                        StatRecords1Norm['Last 3 Win'].iloc[xLast3Norm] + StatRecords1Norm['Last 3 Loss'].iloc[
                    xLast3Norm])) * -1
        elif StatRecords1Norm['Last 3 Win'].iloc[xLast3Norm] == StatRecords1Norm['Last 3 Loss'].iloc[xLast3Norm]:
            StatRecords1Norm['Last 3 Win %'].iloc[xLast3Norm] = 0.5
        # Calculates the win % of the stats
        xLast3Norm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xLast1Norm < len(Compare1Norm['Last 1']):
        try:
            # Access the value at the specified index
            value = Compare1Norm['Last 1'].iloc[xLast1Norm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positivesLast1Norm += 1
            if GameOutcomeNorm == 1:
                StatRecords1Norm['Last 1 Win'].iloc[xLast1Norm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords1Norm['Last 1 Loss'].iloc[xLast1Norm] += 1
        elif value < 0:
            negativesLast1Norm += 1
            if GameOutcomeNorm == 2:
                StatRecords1Norm['Last 1 Win'].iloc[xLast1Norm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords1Norm['Last 1 Loss'].iloc[xLast1Norm] += 1

        if StatRecords1Norm['Last 1 Win'].iloc[xLast1Norm] > StatRecords1Norm['Last 1 Loss'].iloc[xLast1Norm]:
            StatRecords1Norm['Last 1 Win %'].iloc[xLast1Norm] = (StatRecords1Norm['Last 1 Win'].iloc[xLast1Norm]) / (
                        StatRecords1Norm['Last 1 Win'].iloc[xLast1Norm] + StatRecords1Norm['Last 1 Loss'].iloc[
                    xLast1Norm])
        elif StatRecords1Norm['Last 1 Win'].iloc[xLast1Norm] < StatRecords1Norm['Last 1 Loss'].iloc[xLast1Norm]:
            StatRecords1Norm['Last 1 Win %'].iloc[xLast1Norm] = ((StatRecords1Norm['Last 1 Loss'].iloc[xLast1Norm]) / (
                        StatRecords1Norm['Last 1 Win'].iloc[xLast1Norm] + StatRecords1Norm['Last 1 Loss'].iloc[
                    xLast1Norm])) * -1
        elif StatRecords1Norm['Last 1 Win'].iloc[xLast1Norm] == StatRecords1Norm['Last 1 Loss'].iloc[xLast1Norm]:
            StatRecords1Norm['Last 1 Win %'].iloc[xLast1Norm] = 0.5
        # Calculates win % of the stats
        xLast1Norm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xHomeNorm < len(Compare1Norm['Home']):
        try:
            # Access the value at the specified index
            value = Compare1Norm['Home'].iloc[xHomeNorm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positivesHomeNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords1Norm['Home Win'].iloc[xHomeNorm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords1Norm['Home Loss'].iloc[xHomeNorm] += 1
        elif value < 0:
            negativesHomeNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords1Norm['Home Win'].iloc[xHomeNorm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords1Norm['Home Loss'].iloc[xHomeNorm] += 1

        if StatRecords1Norm['Home Win'].iloc[xHomeNorm] > StatRecords1Norm['Home Loss'].iloc[xHomeNorm]:
            StatRecords1Norm['Home Win %'].iloc[xHomeNorm] = (StatRecords1Norm['Home Win'].iloc[xHomeNorm]) / (
                        StatRecords1Norm['Home Win'].iloc[xHomeNorm] + StatRecords1Norm['Home Loss'].iloc[xHomeNorm])
        elif StatRecords1Norm['Home Win'].iloc[xHomeNorm] < StatRecords1Norm['Home Loss'].iloc[xHomeNorm]:
            StatRecords1Norm['Home Win %'].iloc[xHomeNorm] = ((StatRecords1Norm['Home Loss'].iloc[xHomeNorm]) / (
                        StatRecords1Norm['Home Win'].iloc[xHomeNorm] + StatRecords1Norm['Home Loss'].iloc[
                    xHomeNorm])) * -1
        elif StatRecords1Norm['Home Win'].iloc[xHomeNorm] == StatRecords1Norm['Home Loss'].iloc[xHomeNorm]:
            StatRecords1Norm['Home Win %'].iloc[xHomeNorm] = 0.5
        # Calculates win % of the stats
        xHomeNorm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xAwayNorm < len(Compare1Norm['Away']):
        try:
            # Access the value at the specified index
            value = Compare1Norm['Away'].iloc[xAwayNorm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positivesAwayNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords1Norm['Away Win'].iloc[xAwayNorm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords1Norm['Away Loss'].iloc[xAwayNorm] += 1
        elif value < 0:
            negativesAwayNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords1Norm['Away Win'].iloc[xAwayNorm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords1Norm['Away Loss'].iloc[xAwayNorm] += 1

        if StatRecords1Norm['Away Win'].iloc[xAwayNorm] > StatRecords1Norm['Away Loss'].iloc[xAwayNorm]:
            StatRecords1Norm['Away Win %'].iloc[xAwayNorm] = (StatRecords1Norm['Away Win'].iloc[xAwayNorm]) / (
                        StatRecords1Norm['Away Win'].iloc[xAwayNorm] + StatRecords1Norm['Away Loss'].iloc[xAwayNorm])
        elif StatRecords1Norm['Away Win'].iloc[xAwayNorm] < StatRecords1Norm['Away Loss'].iloc[xAwayNorm]:
            StatRecords1Norm['Away Win %'].iloc[xAwayNorm] = ((StatRecords1Norm['Away Loss'].iloc[xAwayNorm]) / (
                        StatRecords1Norm['Away Win'].iloc[xAwayNorm] + StatRecords1Norm['Away Loss'].iloc[
                    xAwayNorm])) * -1
        elif StatRecords1Norm['Away Win'].iloc[xAwayNorm] == StatRecords1Norm['Away Loss'].iloc[xAwayNorm]:
            StatRecords1Norm['Away Win %'].iloc[xAwayNorm] = 0.5
        # Calculates win % of the stats
        xAwayNorm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xRating1Norm < len(Compare2Norm['Rating']):
        try:
            # Access the value at the specified index
            value = Compare2Norm['Rating'].iloc[xRating1Norm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives1ofHiNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords2Norm['Rating Win'].iloc[xRating1Norm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords2Norm['Rating Loss'].iloc[xRating1Norm] += 1
        elif value < 0:
            negatives1ofHiNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords2Norm['Rating Win'].iloc[xRating1Norm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords2Norm['Rating Loss'].iloc[xRating1Norm] += 1

        if StatRecords2Norm['Rating Win'].iloc[xRating1Norm] > StatRecords2Norm['Rating Loss'].iloc[xRating1Norm]:
            StatRecords2Norm['Rating Win %'].iloc[xRating1Norm] = (
                                                                  StatRecords2Norm['Rating Win'].iloc[xRating1Norm]) / (
                                                                              StatRecords2Norm['Rating Win'].iloc[
                                                                                  xRating1Norm] +
                                                                              StatRecords2Norm['Rating Loss'].iloc[
                                                                                  xRating1Norm])
        elif StatRecords2Norm['Rating Win'].iloc[xRating1Norm] < StatRecords2Norm['Rating Loss'].iloc[xRating1Norm]:
            StatRecords2Norm['Rating Win %'].iloc[xRating1Norm] = ((StatRecords2Norm['Rating Loss'].iloc[
                xRating1Norm]) / (StatRecords2Norm['Rating Win'].iloc[xRating1Norm] +
                                  StatRecords2Norm['Rating Loss'].iloc[xRating1Norm])) * -1
        elif StatRecords2Norm['Rating Win'].iloc[xRating1Norm] == StatRecords2Norm['Rating Loss'].iloc[xRating1Norm]:
            StatRecords2Norm['Rating Win %'].iloc[xRating1Norm] = 0.5
        # Calculates win % of the stats
        xRating1Norm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xof1to25Norm < len(Compare2Norm['v 1-25']):
        try:
            # Access the value at the specified index
            value = Compare2Norm['v 1-25'].iloc[xof1to25Norm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives1to25Norm += 1
            if GameOutcomeNorm == 1:
                StatRecords2Norm['v 1-25 Win'].iloc[xof1to25Norm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords2Norm['v 1-25 Loss'].iloc[xof1to25Norm] += 1
        elif value < 0:
            negatives1to25Norm += 1
            if GameOutcomeNorm == 2:
                StatRecords2Norm['v 1-25 Win'].iloc[xof1to25Norm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords2Norm['v 1-25 Loss'].iloc[xof1to25Norm] += 1

        if StatRecords2Norm['v 1-25 Win'].iloc[xof1to25Norm] > StatRecords2Norm['v 1-25 Loss'].iloc[xof1to25Norm]:
            StatRecords2Norm['v 1-25 Win %'].iloc[xof1to25Norm] = (
                                                                  StatRecords2Norm['v 1-25 Win'].iloc[xof1to25Norm]) / (
                                                                              StatRecords2Norm['v 1-25 Win'].iloc[
                                                                                  xof1to25Norm] +
                                                                              StatRecords2Norm['v 1-25 Loss'].iloc[
                                                                                  xof1to25Norm])
        elif StatRecords2Norm['v 1-25 Win'].iloc[xof1to25Norm] < StatRecords2Norm['v 1-25 Loss'].iloc[xof1to25Norm]:
            StatRecords2Norm['v 1-25 Win %'].iloc[xof1to25Norm] = ((StatRecords2Norm['v 1-25 Loss'].iloc[
                xof1to25Norm]) / (StatRecords2Norm['v 1-25 Win'].iloc[xof1to25Norm] +
                                  StatRecords2Norm['v 1-25 Loss'].iloc[xof1to25Norm])) * -1
        elif StatRecords2Norm['v 1-25 Win'].iloc[xof1to25Norm] == StatRecords2Norm['v 1-25 Loss'].iloc[xof1to25Norm]:
            StatRecords2Norm['v 1-25 Win %'].iloc[xof1to25Norm] = 0.5
        # Calculates win % of the stats
        xof1to25Norm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xof26to50Norm < len(Compare2Norm['v 26-50']):
        try:
            # Access the value at the specified index
            value = Compare2Norm['v 26-50'].iloc[xof26to50Norm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives26to50Norm += 1
            if GameOutcomeNorm == 1:
                StatRecords2Norm['v 26-50 Win'].iloc[xof26to50Norm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords2Norm['v 26-50 Loss'].iloc[xof26to50Norm] += 1
        elif value < 0:
            negatives26to50Norm += 1
            if GameOutcomeNorm == 2:
                StatRecords2Norm['v 26-50 Win'].iloc[xof26to50Norm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords2Norm['v 26-50 Loss'].iloc[xof26to50Norm] += 1

        if StatRecords2Norm['v 26-50 Win'].iloc[xof26to50Norm] > StatRecords2Norm['v 26-50 Loss'].iloc[xof26to50Norm]:
            StatRecords2Norm['v 26-50 Win %'].iloc[xof26to50Norm] = (StatRecords2Norm['v 26-50 Win'].iloc[
                xof26to50Norm]) / (StatRecords2Norm['v 26-50 Win'].iloc[xof26to50Norm] +
                                   StatRecords2Norm['v 26-50 Loss'].iloc[xof26to50Norm])
        elif StatRecords2Norm['v 26-50 Win'].iloc[xof26to50Norm] < StatRecords2Norm['v 26-50 Loss'].iloc[xof26to50Norm]:
            StatRecords2Norm['v 26-50 Win %'].iloc[xof26to50Norm] = ((StatRecords2Norm['v 26-50 Loss'].iloc[
                xof26to50Norm]) / (StatRecords2Norm['v 26-50 Win'].iloc[xof26to50Norm] +
                                   StatRecords2Norm['v 26-50 Loss'].iloc[xof26to50Norm])) * -1
        elif StatRecords2Norm['v 26-50 Win'].iloc[xof26to50Norm] == StatRecords2Norm['v 26-50 Loss'].iloc[
            xof26to50Norm]:
            StatRecords2Norm['v 26-50 Win %'].iloc[xof26to50Norm] = 0.5
        # Calculates win % of the stats
        xof26to50Norm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xof51to100Norm < len(Compare2Norm['v 51-100']):
        try:
            # Access the value at the specified index
            value = Compare2Norm['v 51-100'].iloc[xof51to100Norm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives51to100Norm += 1
            if GameOutcomeNorm == 1:
                StatRecords2Norm['v 51-100 Win'].iloc[xof51to100Norm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords2Norm['v 51-100 Loss'].iloc[xof51to100Norm] += 1
        elif value < 0:
            negatives51to100Norm += 1
            if GameOutcomeNorm == 2:
                StatRecords2Norm['v 51-100 Win'].iloc[xof51to100Norm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords2Norm['v 51-100 Loss'].iloc[xof51to100Norm] += 1

        if StatRecords2Norm['v 51-100 Win'].iloc[xof51to100Norm] > StatRecords2Norm['v 51-100 Loss'].iloc[
            xof51to100Norm]:
            StatRecords2Norm['v 51-100 Win %'].iloc[xof51to100Norm] = (StatRecords2Norm['v 51-100 Win'].iloc[
                xof51to100Norm]) / (StatRecords2Norm['v 51-100 Win'].iloc[xof51to100Norm] +
                                    StatRecords2Norm['v 51-100 Loss'].iloc[xof51to100Norm])
        elif StatRecords2Norm['v 51-100 Win'].iloc[xof51to100Norm] < StatRecords2Norm['v 51-100 Loss'].iloc[
            xof51to100Norm]:
            StatRecords2Norm['v 51-100 Win %'].iloc[xof51to100Norm] = ((StatRecords2Norm['v 51-100 Loss'].iloc[
                xof51to100Norm]) / (StatRecords2Norm['v 51-100 Win'].iloc[xof51to100Norm] +
                                    StatRecords2Norm['v 51-100 Loss'].iloc[xof51to100Norm])) * -1
        elif StatRecords2Norm['v 51-100 Win'].iloc[xof51to100Norm] == StatRecords2Norm['v 51-100 Loss'].iloc[
            xof51to100Norm]:
            StatRecords2Norm['v 51-100 Win %'].iloc[xof51to100Norm] = 0.5
        # Calculates win % of the stats
        xof51to100Norm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while x1ofHiNorm < len(Compare2Norm['Hi']):
        try:
            # Access the value at the specified index
            value = Compare2Norm['Hi'].iloc[x1ofHiNorm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives1ofHiNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords2Norm['Hi Win'].iloc[x1ofHiNorm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords2Norm['Hi Loss'].iloc[x1ofHiNorm] += 1
        elif value < 0:
            negatives1ofHiNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords2Norm['Hi Win'].iloc[x1ofHiNorm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords2Norm['Hi Loss'].iloc[x1ofHiNorm] += 1

        if StatRecords2Norm['Hi Win'].iloc[x1ofHiNorm] > StatRecords2Norm['Hi Loss'].iloc[x1ofHiNorm]:
            StatRecords2Norm['Hi Win %'].iloc[x1ofHiNorm] = (StatRecords2Norm['Hi Win'].iloc[x1ofHiNorm]) / (
                        StatRecords2Norm['Hi Win'].iloc[x1ofHiNorm] + StatRecords2Norm['Hi Loss'].iloc[x1ofHiNorm])
        elif StatRecords2Norm['Hi Win'].iloc[x1ofHiNorm] < StatRecords2Norm['Hi Loss'].iloc[x1ofHiNorm]:
            StatRecords2Norm['Hi Win %'].iloc[x1ofHiNorm] = ((StatRecords2Norm['Hi Loss'].iloc[x1ofHiNorm]) / (
                        StatRecords2Norm['Hi Win'].iloc[x1ofHiNorm] + StatRecords2Norm['Hi Loss'].iloc[
                    x1ofHiNorm])) * -1
        elif StatRecords2Norm['Hi Win'].iloc[x1ofHiNorm] == StatRecords2Norm['Hi Loss'].iloc[x1ofHiNorm]:
            StatRecords2Norm['Hi Win %'].iloc[x1ofHiNorm] = 0.5
        # Calculates win % of the stats
        x1ofHiNorm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while x1ofLowNorm < len(Compare2Norm['Low']):
        try:
            # Access the value at the specified index
            value = Compare2Norm['Low'].iloc[x1ofLowNorm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives1ofLowNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords2Norm['Low Win'].iloc[x1ofLowNorm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords2Norm['Low Loss'].iloc[x1ofLowNorm] += 1
        elif value < 0:
            negatives1ofLowNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords2Norm['Low Win'].iloc[x1ofLowNorm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords2Norm['Low Loss'].iloc[x1ofLowNorm] += 1

        if StatRecords2Norm['Low Win'].iloc[x1ofLowNorm] > StatRecords2Norm['Low Loss'].iloc[x1ofLowNorm]:
            StatRecords2Norm['Low Win %'].iloc[x1ofLowNorm] = (StatRecords2Norm['Low Win'].iloc[x1ofLowNorm]) / (
                        StatRecords2Norm['Low Win'].iloc[x1ofLowNorm] + StatRecords2Norm['Low Loss'].iloc[x1ofLowNorm])
        elif StatRecords2Norm['Low Win'].iloc[x1ofLowNorm] < StatRecords2Norm['Low Loss'].iloc[x1ofLowNorm]:
            StatRecords2Norm['Low Win %'].iloc[x1ofLowNorm] = ((StatRecords2Norm['Low Loss'].iloc[x1ofLowNorm]) / (
                        StatRecords2Norm['Low Win'].iloc[x1ofLowNorm] + StatRecords2Norm['Low Loss'].iloc[
                    x1ofLowNorm])) * -1
        elif StatRecords2Norm['Low Win'].iloc[x1ofLowNorm] == StatRecords2Norm['Low Loss'].iloc[x1ofLowNorm]:
            StatRecords2Norm['Low Win %'].iloc[x1ofLowNorm] = 0.5
        # Calculates win % of the stats
        x1ofLowNorm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while x1ofLastNorm < len(Compare2Norm['Last']):
        try:
            # Access the value at the specified index
            value = Compare2Norm['Last'].iloc[x1ofLastNorm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives1ofLastNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords2Norm['Last Win'].iloc[x1ofLastNorm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords2Norm['Last Loss'].iloc[x1ofLastNorm] += 1
        elif value < 0:
            negatives1ofLastNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords2Norm['Last Win'].iloc[x1ofLastNorm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords2Norm['Last Loss'].iloc[x1ofLastNorm] += 1

        if StatRecords2Norm['Last Win'].iloc[x1ofLastNorm] > StatRecords2Norm['Last Loss'].iloc[x1ofLastNorm]:
            StatRecords2Norm['Last Win %'].iloc[x1ofLastNorm] = (StatRecords2Norm['Last Win'].iloc[x1ofLastNorm]) / (
                        StatRecords2Norm['Last Win'].iloc[x1ofLastNorm] + StatRecords2Norm['Last Loss'].iloc[
                    x1ofLastNorm])
        elif StatRecords2Norm['Last Win'].iloc[x1ofLastNorm] < StatRecords2Norm['Last Loss'].iloc[x1ofLastNorm]:
            StatRecords2Norm['Last Win %'].iloc[x1ofLastNorm] = ((StatRecords2Norm['Last Loss'].iloc[x1ofLastNorm]) / (
                        StatRecords2Norm['Last Win'].iloc[x1ofLastNorm] + StatRecords2Norm['Last Loss'].iloc[
                    x1ofLastNorm])) * -1
        elif StatRecords2Norm['Last Win'].iloc[x1ofLastNorm] == StatRecords2Norm['Last Loss'].iloc[x1ofLastNorm]:
            StatRecords2Norm['Last Win %'].iloc[x1ofLastNorm] = 0.5
        # Calculates win % of the stats
        x1ofLastNorm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while xRating2Norm < len(Compare3Norm['Rating']):
        try:
            # Access the value at the specified index
            value = Compare3Norm['Rating'].iloc[xRating2Norm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives2ofHiNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords3Norm['Rating Win'].iloc[xRating2Norm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords3Norm['Rating Loss'].iloc[xRating2Norm] += 1
        elif value < 0:
            negatives2ofHiNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords3Norm['Rating Win'].iloc[xRating2Norm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords3Norm['Rating Loss'].iloc[xRating2Norm] += 1

        if StatRecords3Norm['Rating Win'].iloc[xRating2Norm] > StatRecords3Norm['Rating Loss'].iloc[xRating2Norm]:
            StatRecords3Norm['Rating Win %'].iloc[xRating2Norm] = (
                                                                  StatRecords3Norm['Rating Win'].iloc[xRating2Norm]) / (
                                                                              StatRecords3Norm['Rating Win'].iloc[
                                                                                  xRating2Norm] +
                                                                              StatRecords3Norm['Rating Loss'].iloc[
                                                                                  xRating2Norm])
        elif StatRecords3Norm['Rating Win'].iloc[xRating2Norm] < StatRecords3Norm['Rating Loss'].iloc[xRating2Norm]:
            StatRecords3Norm['Rating Win %'].iloc[xRating2Norm] = ((StatRecords3Norm['Rating Loss'].iloc[
                xRating2Norm]) / (StatRecords3Norm['Rating Win'].iloc[xRating2Norm] +
                                  StatRecords3Norm['Rating Loss'].iloc[xRating2Norm])) * -1
        elif StatRecords3Norm['Rating Win'].iloc[xRating2Norm] == StatRecords3Norm['Rating Loss'].iloc[xRating2Norm]:
            StatRecords3Norm['Rating Win %'].iloc[xRating2Norm] = 0.5
        # Calculates win % of the stats
        xRating2Norm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while x2ofHiNorm < len(Compare3Norm['Hi']):
        try:
            # Access the value at the specified index
            value = Compare3Norm['Hi'].iloc[x2ofHiNorm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives2ofHiNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords3Norm['Hi Win'].iloc[x2ofHiNorm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords3Norm['Hi Loss'].iloc[x2ofHiNorm] += 1
        elif value < 0:
            negatives2ofHiNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords3Norm['Hi Win'].iloc[x2ofHiNorm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords3Norm['Hi Loss'].iloc[x2ofHiNorm] += 1

        if StatRecords3Norm['Hi Win'].iloc[x2ofHiNorm] > StatRecords3Norm['Hi Loss'].iloc[x2ofHiNorm]:
            StatRecords3Norm['Hi Win %'].iloc[x2ofHiNorm] = (StatRecords3Norm['Hi Win'].iloc[x2ofHiNorm]) / (
                        StatRecords3Norm['Hi Win'].iloc[x2ofHiNorm] + StatRecords3Norm['Hi Loss'].iloc[x2ofHiNorm])
        elif StatRecords3Norm['Hi Win'].iloc[x2ofHiNorm] < StatRecords3Norm['Hi Loss'].iloc[x2ofHiNorm]:
            StatRecords3Norm['Hi Win %'].iloc[x2ofHiNorm] = ((StatRecords3Norm['Hi Loss'].iloc[x2ofHiNorm]) / (
                        StatRecords3Norm['Hi Win'].iloc[x2ofHiNorm] + StatRecords3Norm['Hi Loss'].iloc[
                    x2ofHiNorm])) * -1
        elif StatRecords3Norm['Hi Win'].iloc[x2ofHiNorm] == StatRecords3Norm['Hi Loss'].iloc[x2ofHiNorm]:
            StatRecords3Norm['Hi Win %'].iloc[x2ofHiNorm] = 0.5
        # Calculates win % of the stats
        x2ofHiNorm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while x2ofLowNorm < len(Compare3Norm['Low']):
        try:
            # Access the value at the specified index
            value = Compare3Norm['Low'].iloc[x2ofLowNorm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives2ofLowNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords3Norm['Low Win'].iloc[x2ofLowNorm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords3Norm['Low Loss'].iloc[x2ofLowNorm] += 1
        elif value < 0:
            negatives2ofLowNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords3Norm['Low Win'].iloc[x2ofLowNorm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords3Norm['Low Loss'].iloc[x2ofLowNorm] += 1

        if StatRecords3Norm['Low Win'].iloc[x2ofLowNorm] > StatRecords3Norm['Low Loss'].iloc[x2ofLowNorm]:
            StatRecords3Norm['Low Win %'].iloc[x2ofLowNorm] = (StatRecords3Norm['Low Win'].iloc[x2ofLowNorm]) / (
                        StatRecords3Norm['Low Win'].iloc[x2ofLowNorm] + StatRecords3Norm['Low Loss'].iloc[x2ofLowNorm])
        elif StatRecords3Norm['Low Win'].iloc[x2ofLowNorm] < StatRecords3Norm['Low Loss'].iloc[x2ofLowNorm]:
            StatRecords3Norm['Low Win %'].iloc[x2ofLowNorm] = ((StatRecords3Norm['Low Loss'].iloc[x2ofLowNorm]) / (
                        StatRecords3Norm['Low Win'].iloc[x2ofLowNorm] + StatRecords3Norm['Low Loss'].iloc[
                    x2ofLowNorm])) * -1
        elif StatRecords3Norm['Low Win'].iloc[x2ofLowNorm] == StatRecords3Norm['Low Loss'].iloc[x2ofLowNorm]:
            StatRecords3Norm['Low Win %'].iloc[x2ofLowNorm] = 0.5
        # Calculates win % of the stats
        x2ofLowNorm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    while x2ofLastNorm < len(Compare3Norm['Last']):
        try:
            # Access the value at the specified index
            value = Compare3Norm['Last'].iloc[x2ofLastNorm]
        except IndexError:
            # Handle the error by setting the value to 0
            value = 0
        # Continue with the rest of the code
        if value > 0:
            positives2ofLastNorm += 1
            if GameOutcomeNorm == 1:
                StatRecords3Norm['Last Win'].iloc[x2ofLastNorm] += 1
            elif GameOutcomeNorm == 2:
                StatRecords3Norm['Last Loss'].iloc[x2ofLastNorm] += 1
        elif value < 0:
            negatives2ofLastNorm += 1
            if GameOutcomeNorm == 2:
                StatRecords3Norm['Last Win'].iloc[x2ofLastNorm] += 1
            elif GameOutcomeNorm == 1:
                StatRecords3Norm['Last Loss'].iloc[x2ofLastNorm] += 1

        if StatRecords3Norm['Last Win'].iloc[x2ofLastNorm] > StatRecords3Norm['Last Loss'].iloc[x2ofLastNorm]:
            StatRecords3Norm['Last Win %'].iloc[x2ofLastNorm] = (StatRecords3Norm['Last Win'].iloc[x2ofLastNorm]) / (
                        StatRecords3Norm['Last Win'].iloc[x2ofLastNorm] + StatRecords3Norm['Last Loss'].iloc[
                    x2ofLastNorm])
        elif StatRecords3Norm['Last Win'].iloc[x2ofLastNorm] < StatRecords3Norm['Last Loss'].iloc[x2ofLastNorm]:
            StatRecords3Norm['Last Win %'].iloc[x2ofLastNorm] = ((StatRecords3Norm['Last Loss'].iloc[x2ofLastNorm]) / (
                        StatRecords3Norm['Last Win'].iloc[x2ofLastNorm] + StatRecords3Norm['Last Loss'].iloc[
                    x2ofLastNorm])) * -1
        elif StatRecords3Norm['Last Win'].iloc[x2ofLastNorm] == StatRecords3Norm['Last Loss'].iloc[x2ofLastNorm]:
            StatRecords3Norm['Last Win %'].iloc[x2ofLastNorm] = 0.5
        # Calculates win % of the stats
        x2ofLastNorm += 1
        # Loops through all the values in Column '2022' and counts how many positive and negative values there are
        # Also adds a win or loss to the StatsRecords column

    StatRecords1Norm.to_csv("SR1Norm.csv", index=False)

    StatRecords2Norm.to_csv("SR2Norm.csv", index=False)

    StatRecords3Norm.to_csv("SR3Norm.csv", index=False)

    Compare1NormSum2022 = Compare1Norm['2022'].sum()
    Compare1NormSumLast3 = Compare1Norm['Last 3'].sum()
    Compare1NormSumLast1 = Compare1Norm['Last 1'].sum()
    Compare1NormSumHome = Compare1Norm['Home'].sum()
    Compare1NormSumAway = Compare1Norm['Away'].sum()
    Compare2NormSumRating = Compare2Norm['Rating'].sum()
    Compare2NormSum1to25 = Compare2Norm['v 1-25'].sum()
    Compare2NormSum26to50 = Compare2Norm['v 26-50'].sum()
    Compare2NormSum51to100 = Compare2Norm['v 51-100'].sum()
    Compare2NormSumHi = Compare2Norm['Hi'].sum()
    Compare2NormSumLow = Compare2Norm['Low'].sum()
    Compare2NormSumLast = Compare2Norm['Last'].sum()
    Compare3NormSumRating = Compare3Norm['Rating'].sum()
    Compare3NormSumHi = Compare3Norm['Hi'].sum()
    Compare3NormSumLow = Compare3Norm['Low'].sum()
    Compare3NormSumLast = Compare3Norm['Last'].sum()

    Compare1Norm2022Mult = Compare1Norm['2022'].multiply(StatRecords1Norm['2022 Win %'])
    Compare1NormLast3Mult = Compare1Norm['Last 3'].multiply(StatRecords1Norm['Last 3 Win %'])
    Compare1NormLast1Mult = Compare1Norm['Last 1'].multiply(StatRecords1Norm['Last 1 Win %'])
    Compare1NormHomeMult = Compare1Norm['Home'].multiply(StatRecords1Norm['Home Win %'])
    Compare1NormAwayMult = Compare1Norm['Away'].multiply(StatRecords1Norm['Away Win %'])

    Compare2NormRatingMult = Compare2Norm['Rating'].multiply(StatRecords2Norm['Rating Win %'])
    Compare2Norm1to25Mult = Compare2Norm['v 1-25'].multiply(StatRecords2Norm['v 1-25 Win %'])
    Compare2Norm26to50Mult = Compare2Norm['v 26-50'].multiply(StatRecords2Norm['v 26-50 Win %'])
    Compare2Norm51to100Mult = Compare2Norm['v 51-100'].multiply(StatRecords2Norm['v 51-100 Win %'])
    Compare2NormHiMult = Compare2Norm['Hi'].multiply(StatRecords2Norm['Hi Win %'])
    Compare2NormLowMult = Compare2Norm['Low'].multiply(StatRecords2Norm['Low Win %'])
    Compare2NormLastMult = Compare2Norm['Last'].multiply(StatRecords2Norm['Last Win %'])

    Compare3NormRatingMult = Compare3Norm['Rating'].multiply(StatRecords3Norm['Rating Win %'])
    Compare3NormHiMult = Compare3Norm['Hi'].multiply(StatRecords3Norm['Hi Win %'])
    Compare3NormLowMult = Compare3Norm['Low'].multiply(StatRecords3Norm['Low Win %'])
    Compare3NormLastMult = Compare3Norm['Last'].multiply(StatRecords3Norm['Last Win %'])

    Compare1Norm2022MultSum = Compare1Norm2022Mult.sum()
    Compare1NormLast3MultSum = Compare1NormLast3Mult.sum()
    Compare1NormLast1MultSum = Compare1NormLast1Mult.sum()
    Compare1NormHomeMultSum = Compare1NormHomeMult.sum()
    Compare1NormAwayMultSum = Compare1NormAwayMult.sum()

    Compare2NormRatingMultSum = Compare2NormRatingMult.sum()
    Compare2Norm1to25MultSum = Compare2Norm1to25Mult.sum()
    Compare2Norm26to50MultSum = Compare2Norm26to50Mult.sum()
    Compare2Norm51to100MultSum = Compare2Norm51to100Mult.sum()
    Compare2NormHiMultSum = Compare2NormHiMult.sum()
    Compare2NormLowMultSum = Compare2NormLowMult.sum()
    Compare2NormLastMultSum = Compare2NormLastMult.sum()

    Compare3NormRatingMultSum = Compare3NormRatingMult.sum()
    Compare3NormHiMultSum = Compare3NormHiMult.sum()
    Compare3NormLowMultSum = Compare3NormLowMult.sum()
    Compare3NormLastMultSum = Compare3NormLastMult.sum()

    AllSum = sum(
        [Compare1NormSum2022, Compare1NormSumLast3, Compare1NormSumLast1, Compare1NormSumHome, Compare1NormSumAway,
         Compare2NormSumRating, Compare2NormSum1to25, Compare2NormSum26to50, Compare2NormSum51to100, Compare2NormSumHi,
         Compare2NormSumLow, Compare2NormSumLast, Compare3NormSumRating, Compare3NormSumHi, Compare3NormSumLow,
         Compare3NormSumLast])

    AllSumMult = sum([Compare1Norm2022MultSum,
                     Compare1NormLast3MultSum,
                     Compare1NormLast1MultSum,
                     Compare1NormHomeMultSum,
                     Compare1NormAwayMultSum,
                     Compare2NormRatingMultSum,
                     Compare2Norm1to25MultSum,
                     Compare2Norm26to50MultSum,
                     Compare2Norm51to100MultSum,
                     Compare2NormHiMultSum,
                     Compare2NormLowMultSum,
                     Compare2NormLastMultSum,
                     Compare3NormRatingMultSum,
                     Compare3NormHiMultSum,
                     Compare3NormLowMultSum,
                     Compare3NormLastMultSum])

    # Open the workbook and get the sheet
    workbook = openpyxl.load_workbook('outputexcel.xlsx')
    sheet = workbook.worksheets[0]

    # Append the variables to the sheet
    sheet.append([Team1, Team2, AllSum, AllSumMult])

    # Save the workbook
    workbook.save('outputexcel.xlsx')

    print(Team1)
    print(Team2)
    print(AllSum)
    print(AllSumMult)
    print(GameOutcomeNorm)
    print(Compare1Norm)
    print(Compare2Norm)
    print(Compare3Norm)

    Compare1Norm.to_csv("Compare1Norm.csv", index=False)

    Compare2Norm.to_csv("Compare2Norm.csv", index=False)

    Compare3Norm.to_csv("Compare3Norm.csv", index=False)



