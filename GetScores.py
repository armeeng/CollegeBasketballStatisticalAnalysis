import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import os
import csv
import openpyxl
from openpyxl import Workbook

# Create a workbook and add a new sheet
workbook = openpyxl.load_workbook('outputexcel.xlsx')
sheet = workbook.worksheets[2]
sheet1 = workbook.worksheets[0]
sheet.delete_rows(1, sheet.max_row)

# Send a GET request to the ESPN scoreboard page
url = "https://www.espn.com/mens-college-basketball/scoreboard/_/seasontype/2/group/50"
response = requests.get(url)

# Parse the HTML content of the page
soup = BeautifulSoup(response.content, "html.parser")

# Find all the score cells
score_cells = soup.find_all("li", class_="ScoreboardScoreCell__Item")

# Extract the scores from each score cell
for score_cell in score_cells:
    # Find the away team name and score elements
    away_team_name_elem = score_cell.find("div", class_="ScoreCell__TeamName")
    away_score_elem = score_cell.find("div", class_="ScoreCell__Score")

    # Extract the text from the elements
    away_team_name = away_team_name_elem.text if away_team_name_elem is not None else "N/A"
    away_score = away_score_elem.text if away_score_elem is not None else "N/A"
    if away_score == "N/A":
        away_score = 0

    #print(f"{away_team_name}: {away_score}")
    # Add the variables as the first row of the sheet
    sheet.append([away_team_name, away_score])

    # Save the workbook
    workbook.save('outputexcel.xlsx')

workbook = openpyxl.load_workbook('outputexcel.xlsx')
sheet1 = workbook.worksheets[0]
sheet2 = workbook.worksheets[2]

TeamRankingsName = ["Abl Christian",
"Air Force",
"Akron",
"Alab A&M",
"Alabama",
"Alabama St",
"Albany",
"Alcorn State",
"American",
"App State",
"Arizona St",
"Arizona",
"Arkansas",
"Arkansas St",
"Ark Pine Bl",
"Army",
"Auburn",
"Austin Peay",
"BYU",
"Ball State",
"Baylor",
"Bellarmine",
"Belmont",
"Beth-Cook",
"Binghamton",
"Boise State",
"Boston Col",
"Boston U",
"Bowling Grn",
"Bradley",
"Brown",
"Bryant",
"Bucknell",
"Buffalo",
"Butler",
"Cal Poly",
"CS Bakersfld",
"CS Fullerton",
"Cal St Nrdge",
"Cal Baptist",
"California",
"Campbell",
"Canisius",
"Central Ark",
"Central Conn",
"Central Mich",
"Col Charlestn",
"Charl South",
"Charlotte",
"Chattanooga",
"Chicago St",
"Cincinnati",
"Clemson",
"Cleveland St",
"Coastal Car",
"Colgate",
"Colorado",
"Colorado St",
"Columbia",
"Coppin State",
"Cornell",
"Creighton",
"Dartmouth",
"Davidson",
"Dayton",
"DePaul",
"Delaware",
"Delaware St",
"Denver",
"Detroit",
"Drake",
"Drexel",
"Duke",
"Duquesne",
"E Carolina",
"E Tenn St",
"E Illinois",
"E Kentucky",
"E Michigan",
"E Washingtn",
"Elon",
"Evansville",
"Fairfield",
"F Dickinson",
"Florida A&M",
"Fla Atlantic",
"Florida",
"Fla Gulf Cst",
"Florida Intl",
"Florida St",
"Fordham",
"Fresno St",
"Furman",
"Gard-Webb",
"Geo Mason",
"Geo Wshgtn",
"Georgetown",
"Georgia",
"GA Southern",
"Georgia St",
"GA Tech",
"Gonzaga",
"Grambling St",
"Grd Canyon",
"WI-Grn Bay",
"Hampton",
"Hartford",
"Harvard",
"Hawaii",
"High Point",
"Hofstra",
"Holy Cross",
"Hsn Christian",
"Houston",
"Howard",
"IUPUI",
"Idaho State",
"Idaho",
"Illinois",
"Illinois St",
"Incar Word",
"Indiana",
"Indiana St",
"Iona",
"Iowa",
"Iowa State",
"Jackson St",
"Jacksonville",
"Jksnville St",
"James Mad",
"UMKC",
"Kansas",
"Kansas St",
"Kennesaw St",
"Kent State",
"Kentucky",
"LSU",
"La Salle",
"Lafayette",
"Lamar",
"Lehigh",
"Liberty",
"Lindenwood",
"Lipscomb",
"AR Lit Rock",
"Lg Beach St",
"LIU",
"Longwood",
"LA Lafayette",
"LA Tech",
"Louisville",
"Loyola-Chi",
"Loyola-MD",
"Loyola Mymt",
"Maine",
"Manhattan",
"Marist",
"Marquette",
"Marshall",
"Maryland",
"Maryland ES",
"McNeese St",
"Memphis",
"Mercer",
"Merrimack",
"Miami (OH)",
"Miami (FL)",
"Michigan St",
"Michigan",
"Middle Tenn",
"WI-Milwkee",
"Minnesota",
"Miss State",
"Miss Val St",
"Missouri St",
"Missouri",
"Monmouth",
"Montana",
"Montana St",
"Morehead St",
"Morgan St",
"Mt St Marys",
"Murray St",
"NC State",
"NJIT",
"Navy",
"Nebraska",
"Nevada",
"N Hampshire",
"New Mexico",
"N Mex State",
"New Orleans",
"Niagara",
"Nicholls St",
"Norfolk St",
"N Alabama",
"NC A&T",
"NC Central",
"N Carolina",
"North Dakota",
"N Dakota St",
"N Florida",
"North Texas",
"Northeastrn",
"N Arizona",
"N Colorado",
"N Illinois",
"N Iowa",
"N Kentucky",
"NW State",
"Northwestern",
"Notre Dame",
"Oakland",
"Ohio",
"Ohio State",
"Oklahoma",
"Oklahoma St",
"Old Dominion",
"Mississippi",
"Neb Omaha",
"Oral Roberts",
"Oregon",
"Oregon St",
"Pacific",
"Penn State",
"U Penn",
"Pepperdine",
"Pittsburgh",
"Portland",
"Portland St",
"Prairie View",
"Presbyterian",
"Princeton",
"Providence",
"Purdue",
"IPFW",
"Queens",
"Quinnipiac",
"Radford",
"Rhode Island",
"Rice",
"Richmond",
"Rider",
"Rob Morris",
"Rutgers",
"SE Louisiana",
"SIU Edward",
"S Methodist",
"Sac State",
"Sacred Hrt",
"St Josephs",
"Saint Louis",
"St Marys",
"St Peters",
"Sam Hous St",
"Samford",
"San Diego St",
"San Diego",
"San Francisco",
"San Jose St",
"Santa Clara",
"Seattle",
"Seton Hall",
"Siena",
"S Alabama",
"S Carolina",
"S Car State",
"SC Upstate",
"South Dakota",
"S Dakota St",
"S Florida",
"SE Missouri",
"S Illinois",
"S Indiana",
"Southern",
"S Mississippi",
"S Utah",
"St Bonavent",
"St Fran (PA)",
"St Fran (NY)",
"St Johns",
"St. Thomas",
"Stanford",
"Ste F Austin",
"Stetson",
"Stonehill",
"Stony Brook",
"Syracuse",
"TX Christian",
"Tarleton State",
"Temple",
"TN State",
"TN Tech",
"Tennessee",
"Texas A&M",
"TX A&M-Com",
"TX A&M-CC",
"Texas",
"TX Southern",
"Texas State",
"Texas Tech",
"Citadel",
"Toledo",
"Towson",
"Troy",
"Tulane",
"Tulsa",
"UAB",
"UC Davis",
"UC Irvine",
"UC Riverside",
"UC San Diego",
"UCSB",
"Central FL",
"UCLA",
"Connecticut",
"IL-Chicago",
"LA Monroe",
"Maryland BC",
"Mass Lowell",
"U Mass",
"NC-Asheville",
"NC-Grnsboro",
"NC-Wilmgton",
"UNLV",
"USC",
"TX-Arlington",
"TN Martin",
"TX-Pan Am",
"TX El Paso",
"TX-San Ant",
"Utah State",
"Utah Tech",
"Utah",
"Utah Val St",
"VCU",
"VA Military",
"Valparaiso",
"Vanderbilt",
"Vermont",
"Villanova",
"Virginia",
"VA Tech",
"Wagner",
"Wake Forest",
"Washington",
"Wash State",
"Weber State",
"W Virginia",
"W Carolina",
"W Illinois",
"W Kentucky",
"W Michigan",
"Wichita St",
"Wm & Mary",
"Winthrop",
"Wisconsin",
"Wofford",
"Wright State",
"Wyoming",
"Xavier",
"Yale",
"Youngs St"]
ESPNNames = ["Abilene Christian",
"Air Force",
"Akron",
"Alabama A&M",
"Alabama",
"Alabama State",
"Albany",
"Alcorn State",
"American University",
"Appalachian State",
"Arizona State",
"Arizona",
"Arkansas",
"Arkansas State",
"Arkansas-Pine Bluff",
"Army",
"Auburn",
"Austin Peay",
"BYU",
"Ball State",
"Baylor",
"Bellarmine",
"Belmont",
"Bethune-Cookman",
"Binghamton",
"Boise State",
"Boston College",
"Boston University",
"Bowling Green",
"Bradley",
"Brown",
"Bryant",
"Bucknell",
"Buffalo",
"Butler",
"Cal Poly",
"Cal State Bakersfield",
"Cal State Fullerton",
"Cal State Northridge",
"California Baptist",
"California",
"Campbell",
"Canisius",
"Central Arkansas",
"Central Connecticut",
"Central Michigan",
"Charleston",
"Charleston Southern",
"Charlotte",
"Chattanooga",
"Chicago State",
"Cincinnati",
"Clemson",
"Cleveland State",
"Coastal Carolina",
"Colgate",
"Colorado",
"Colorado State",
"Columbia",
"Coppin State",
"Cornell",
"Creighton",
"Dartmouth",
"Davidson",
"Dayton",
"DePaul",
"Delaware",
"Delaware State",
"Denver",
"Detroit",
"Drake",
"Drexel",
"Duke",
"Duquesne",
"East Carolina",
"East Tennessee State",
"Eastern Illinois",
"Eastern Kentucky",
"Eastern Michigan",
"Eastern Washington",
"Elon",
"Evansville",
"Fairfield",
"Fairleigh Dickinson",
"Florida A&M",
"Florida Atlantic",
"Florida",
"Florida Gulf Coast",
"Florida International",
"Florida State",
"Fordham",
"Fresno State",
"Furman",
"Gardner-Webb",
"George Mason",
"George Washington",
"Georgetown",
"Georgia",
"Georgia Southern",
"Georgia State",
"Georgia Tech",
"Gonzaga",
"Grambling",
"Grand Canyon",
"Green Bay",
"Hampton",
"Hartford",
"Harvard",
"Hawai'i",
"High Point",
"Hofstra",
"Holy Cross",
"Houston Christian",
"Houston",
"Howard",
"IUPUI",
"Idaho State",
"Idaho",
"Illinois",
"Illinois State",
"Incarnate Word",
"Indiana",
"Indiana State",
"Iona",
"Iowa",
"Iowa State",
"Jackson State",
"Jacksonville",
"Jacksonville State",
"James Madison",
"Kansas City",
"Kansas",
"Kansas State",
"Kennesaw State",
"Kent State",
"Kentucky",
"LSU",
"La Salle",
"Lafayette",
"Lamar",
"Lehigh",
"Liberty",
"Lindenwood",
"Lipscomb",
"Little Rock",
"Long Beach State",
"Long Island University",
"Longwood",
"Louisiana",
"Louisiana Tech",
"Louisville",
"Loyola Chicago",
"Loyola Maryland",
"Loyola Marymount",
"Maine",
"Manhattan",
"Marist",
"Marquette",
"Marshall",
"Maryland",
"Maryland-Eastern",
"McNeese",
"Memphis",
"Mercer",
"Merrimack",
"Miami (OH)",
"Miami",
"Michigan State",
"Michigan",
"Middle Tennessee",
"Milwaukee",
"Minnesota",
"Mississippi State",
"Mississippi Valley State",
"Missouri State",
"Missouri",
"Monmouth",
"Montana",
"Montana State",
"Morehead State",
"Morgan State",
"Mount St. Mary's",
"Murray State",
"NC State",
"NJIT",
"Navy",
"Nebraska",
"Nevada",
"New Hampshire",
"New Mexico",
"New Mexico State",
"New Orleans",
"Niagara",
"Nicholls",
"Norfolk State",
"North Alabama",
"North Carolina A&T",
"North Carolina Central",
"North Carolina",
"North Dakota",
"North Dakota State",
"North Florida",
"North Texas",
"Northeastern",
"Northern Arizona",
"Northern Colorado",
"Northern Illinois",
"Northern Iowa",
"Northern Kentucky",
"Northwestern State",
"Northwestern",
"Notre Dame",
"Oakland",
"Ohio",
"Ohio State",
"Oklahoma",
"Oklahoma State",
"Old Dominion",
"Ole Miss",
"Omaha",
"Oral Roberts",
"Oregon",
"Oregon State",
"Pacific",
"Penn State",
"Pennsylvania",
"Pepperdine",
"Pittsburgh",
"Portland",
"Portland State",
"Prairie View A&M",
"Presbyterian",
"Princeton",
"Providence",
"Purdue",
"Purdue Fort Wayne",
"Queens University",
"Quinnipiac",
"Radford",
"Rhode Island",
"Rice",
"Richmond",
"Rider",
"Robert Morris",
"Rutgers",
"SE Louisiana",
"SIU Edwardsville",
"SMU",
"Sacramento State",
"Sacred Heart",
"Saint Joseph's",
"Saint Louis",
"Saint Mary's",
"Saint Peter's",
"Sam Houston",
"Samford",
"San Diego State",
"San Diego",
"San Francisco",
"San José State",
"Santa Clara",
"Seattle U",
"Seton Hall",
"Siena",
"South Alabama",
"South Carolina",
"South Carolina State",
"South Carolina Upstate",
"South Dakota",
"South Dakota State",
"South Florida",
"Southeast Missouri State",
"Southern Illinois",
"Southern Indiana",
"Southern",
"Southern Miss",
"Southern Utah",
"St. Bonaventure",
"St. Francis (PA)",
"St. Francis Brooklyn",
"St. John's",
"St. Thomas - Minnesota",
"Stanford",
"Stephen F. Austin",
"Stetson",
"Stonehill",
"Stony Brook",
"Syracuse",
"TCU",
"Tarleton",
"Temple",
"Tennessee State",
"Tennessee Tech",
"Tennessee",
"Texas A&M",
"Texas A&M-Commerce",
"Texas A&M-Corpus Christi",
"Texas",
"Texas Southern",
"Texas State",
"Texas Tech",
"The Citadel",
"Toledo",
"Towson",
"Troy",
"Tulane",
"Tulsa",
"UAB",
"UC Davis",
"UC Irvine",
"UC Riverside",
"UC San Diego",
"UC Santa Barbara",
"UCF",
"UCLA",
"UConn",
"UIC",
"UL Monroe",
"UMBC",
"UMass Lowell",
"UMass",
"UNC Asheville",
"UNC Greensboro",
"UNC Wilmington",
"UNLV",
"USC",
"UT Arlington",
"UT Martin",
"UT Rio Grande Valley",
"UTEP",
"UTSA",
"Utah State",
"Utah Tech",
"Utah",
"Utah Valley",
"VCU",
"VMI",
"Valparaiso",
"Vanderbilt",
"Vermont",
"Villanova",
"Virginia",
"Virginia Tech",
"Wagner",
"Wake Forest",
"Washington",
"Washington State",
"Weber State",
"West Virginia",
"Western Carolina",
"Western Illinois",
"Western Kentucky",
"Western Michigan",
"Wichita State",
"William & Mary",
"Winthrop",
"Wisconsin",
"Wofford",
"Wright State",
"Wyoming",
"Xavier",
"Yale",
"Youngstown State"]
row_index = 1

for row in sheet1:
    values = [cell.value for cell in row]
    Team1 = values[0]
    Team2 = values[1]
    index1 = TeamRankingsName.index(Team1)
    index2 = TeamRankingsName.index(Team2)
    Team1 = ESPNNames[index1]
    Team2 = ESPNNames[index2]
    for row in sheet2:
        values = [cell.value for cell in row]
        if Team1 == values[0]:
            Score1 = values[1]
            Score1 = int(Score1)
    for row in sheet2:
        values = [cell.value for cell in row]
        if Team2 == values[0]:
            Score2 =values[1]
            Score2 = int(Score2)
    try:
        sheet1.cell(row=row_index, column=7).value = Team1
        sheet1.cell(row=row_index, column=8).value = Score1
        sheet1.cell(row=row_index, column=9).value = Score2
        sheet1.cell(row=row_index, column=10).value = Team2
    except ValueError:
        pass
    row_index += 1
    workbook.save('outputexcel.xlsx')







